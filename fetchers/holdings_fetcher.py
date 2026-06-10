# fetchers/holdings_fetcher.py
# Downloads monthly portfolio disclosure Excel files directly from AMC websites.
# Parses stock-level holdings (name, ISIN, sector, weight) and stores in DB.
# Currently supports PPFAS. New AMCs can be added to AMC_CONFIG below.

import requests
import openpyxl
import io
import sys
from pathlib import Path
from datetime import date
from sqlalchemy.orm import Session
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from data.database import engine, FundHolding, SectorMap, init_db


# ── AMC configuration ─────────────────────────────────────────────────────────
# Each entry maps a scheme_code to its Excel download URL and parser config.
# To add a new AMC: add its schemes here with the correct URL pattern.
AMC_CONFIG = {
    # PPFAS — Parag Parikh Flexi Cap Fund (Direct Growth)
    122639: {
        "amc":        "PPFAS",
        "scheme":     "Parag Parikh Flexi Cap Fund",
        "url":        "https://amc.ppfas.com/downloads/portfolio-disclosure/{year}/PPFCF_PPFAS_Monthly_Portfolio_Report_{month_name}_{day}_{year}.xlsx",
        "header_row": 4,    # Row where column headers live (1-indexed)
        "data_start":  7,   # Row where actual stock data begins (1-indexed)
        "cols": {
            "name":       1,   # Column B — instrument name
            "isin":       2,   # Column C — ISIN
            "sector":     3,   # Column D — industry/sector
            "quantity":   4,   # Column E
            "mkt_value":  5,   # Column F
            "weight_pct": 6,   # Column G — % to net assets
        },
    },

    # PPFAS — Parag Parikh ELSS Tax Saver Fund (Direct Growth)
    # scheme_code TBD — add after verifying
    # 118989: {
    #     "amc": "PPFAS",
    #     "url": "https://amc.ppfas.com/downloads/portfolio-disclosure/{year}/PPTSF_PPFAS_Monthly_Portfolio_Report_{month_name}_{day}_{year}.xlsx",
    #     ...
    # },
}

# Month-end dates for URL construction
MONTH_END = {
    1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
    7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31,
}
MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


# ── URL builder ───────────────────────────────────────────────────────────────
def build_url(scheme_code: int, year: int, month: int) -> str:
    """Constructs the Excel download URL for a given scheme and month."""
    config    = AMC_CONFIG[scheme_code]
    day       = MONTH_END[month]
    month_name = MONTH_NAMES[month]

    # February: handle leap year
    if month == 2 and year % 4 == 0:
        day = 29

    return config["url"].format(
        year=year,
        month_name=month_name,
        day=day,
    )


# ── Excel downloader ──────────────────────────────────────────────────────────
def download_excel(url: str) -> openpyxl.Workbook | None:
    """Downloads an Excel file from a URL and returns an openpyxl workbook."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://amc.ppfas.com/",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 404:
            print(f"  ⚠️  File not found (404): {url}")
            return None
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        return wb
    except Exception as e:
        print(f"  ⚠️  Download failed: {e}")
        return None


# ── Excel parser ──────────────────────────────────────────────────────────────
def parse_holdings(wb: openpyxl.Workbook, scheme_code: int) -> list[dict]:
    """
    Parses holdings from an Excel workbook using the AMC config for that scheme.
    Returns a list of holding dicts: {isin, stock_name, sector, weight_pct}
    Skips rows that are section headers (no ISIN, no meaningful data).
    """
    config   = AMC_CONFIG[scheme_code]
    cols     = config["cols"]
    ws       = wb.active
    holdings = []

    for row_idx in range(config["data_start"], ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=col_idx).value
               for col_idx in range(1, ws.max_column + 1)]

        # Skip empty rows
        if not any(row):
            continue

        name       = row[cols["name"]]
        isin       = row[cols["isin"]]
        sector     = row[cols["sector"]]
        weight_pct = row[cols["weight_pct"]]

        # Skip section header rows — they have text in name but no ISIN
        if not isin or not isinstance(isin, str) or not isin.startswith("INE"):
            # Also catch foreign ISINs (start with 2-letter country code, not INE)
            if not isin or not isinstance(isin, str) or len(isin) < 10:
                continue

        # Skip rows with no weight
        if weight_pct is None:
            continue

        try:
            weight_pct = float(weight_pct)
        except (ValueError, TypeError):
            continue

        # Convert decimal weight to percentage (0.0794 → 7.94%)
        if weight_pct < 1:
            weight_pct = round(weight_pct * 100, 4)

        holdings.append({
            "isin":       str(isin).strip(),
            "stock_name": str(name).strip() if name else "",
            "sector":     str(sector).strip() if sector else "",
            "weight_pct": weight_pct,
        })

    return holdings


# ── Store holdings in DB ──────────────────────────────────────────────────────
def store_holdings(scheme_code: int, holdings: list[dict],
                   holding_date: date) -> int:
    """
    Inserts holdings into fund_holdings table. Skips duplicates.
    Also upserts sector data into sector_map table.
    Returns count of holdings rows inserted.
    """
    if not holdings:
        return 0

    holding_rows = []
    sector_rows  = []

    for h in holdings:
        holding_rows.append({
            "scheme_code":  scheme_code,
            "holding_date": holding_date,
            "isin":         h["isin"],
            "stock_name":   h["stock_name"],
            "weight_pct":   h["weight_pct"],
        })
        if h["isin"] and h["sector"]:
            sector_rows.append({
                "isin":   h["isin"],
                "name":   h["stock_name"],
                "sector": h["sector"],
                "source": "amfi_disclosure",
            })

    with Session(engine) as session:
        # Insert holdings — skip duplicates
        stmt = sqlite_insert(FundHolding).values(holding_rows).prefix_with("OR IGNORE")
        session.execute(stmt)

        # Upsert sector map — update sector if already exists
        for s in sector_rows:
            existing = session.query(SectorMap).filter_by(isin=s["isin"]).first()
            if existing:
                existing.sector = s["sector"]
                existing.name   = s["name"]
            else:
                session.add(SectorMap(
                    isin=s["isin"], name=s["name"],
                    sector=s["sector"], source=s["source"]
                ))

        session.commit()

    return len(holding_rows)


# ── Main fetch function ───────────────────────────────────────────────────────
def fetch_and_store_holdings(scheme_code: int,
                              year: int = None,
                              month: int = None) -> int:
    """
    Downloads and stores holdings for a scheme for a given year/month.
    Defaults to current month.
    """
    today = date.today()
    year  = year  or today.year
    month = month or today.month

    if scheme_code not in AMC_CONFIG:
        print(f"  ⚠️  No AMC config found for scheme {scheme_code}")
        return 0

    url = build_url(scheme_code, year, month)
    print(f"  Downloading: {url}")

    wb = download_excel(url)
    if wb is None:
        return 0

    holdings = parse_holdings(wb, scheme_code)
    if not holdings:
        print(f"  ⚠️  No holdings parsed from file")
        return 0

    holding_date = date(year, month, MONTH_END[month])
    count = store_holdings(scheme_code, holdings, holding_date)
    print(f"  ✅ {count} holdings stored for scheme {scheme_code}")
    return count


# ── Query helper ──────────────────────────────────────────────────────────────
def get_fund_holdings(scheme_code: int) -> list[dict]:
    """Returns the most recent holdings for a fund, sorted by weight desc."""
    with Session(engine) as session:
        latest = (
            session.query(FundHolding.holding_date)
            .filter_by(scheme_code=scheme_code)
            .order_by(FundHolding.holding_date.desc())
            .first()
        )
        if not latest:
            return []

        rows = (
            session.query(FundHolding)
            .filter_by(scheme_code=scheme_code, holding_date=latest[0])
            .order_by(FundHolding.weight_pct.desc())
            .all()
        )

        return [
            {
                "isin":         h.isin,
                "stock_name":   h.stock_name,
                "weight_pct":   h.weight_pct,
                "holding_date": str(h.holding_date),
            }
            for h in rows
        ]


# ── Test ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()

    print("\n── Fetching Parag Parikh Flexi Cap Fund holdings (April 2026) ──")
    count = fetch_and_store_holdings(122639, year=2026, month=4)

    print("\n── Verifying: top 10 holdings from DB ──")
    holdings = get_fund_holdings(122639)

    if holdings:
        print(f"\nFound {len(holdings)} holdings. Top 10:\n")
        print(f"  {'Stock':<45} {'Weight':>8}  {'ISIN'}")
        print(f"  {'-'*45} {'-'*8}  {'-'*14}")
        for h in holdings[:10]:
            print(f"  {h['stock_name']:<45} {h['weight_pct']:>7.2f}%  {h['isin']}")
    else:
        print("  ⚠️  No holdings found in DB")

    print("\n── Sector map sample ──")
    with Session(engine) as session:
        sectors = session.query(SectorMap).limit(10).all()
        for s in sectors:
            print(f"  {s.isin}  {s.name:<40}  {s.sector}")